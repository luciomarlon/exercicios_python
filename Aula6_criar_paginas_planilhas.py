import openpyxl

#criar a planilha
livro = openpyxl.Workbook()

#visualizar páginas existentes
print(livro.sheetnames)

#criar uma página
livro.create_sheet('Frutas')

#selcionar página
pagina_frutas = livro['Frutas']

#adicionando os dados
pagina_frutas.append(['Fruta', 'Quantidade', 'Preço'])
pagina_frutas.append(['Bananas', '5', 'R$4,29'])
pagina_frutas.append(['Maça', '2', 'R$9,90'])
pagina_frutas.append(['laranja', '12', 'R$8,00'])
pagina_frutas.append(['Pera', '10', 'R$4,99'])

#salvar a planilha
livro.save('Planilha de compras.xlsx')

