import openpyxl
import pandas as pd

#carregar arquivo
livro = openpyxl.load_workbook('Planilha de compras.xlsx')

#selecionar uma página
frutas_pagina = livro['Frutas']

#Imprimir os dados de cada linha
for linhas in frutas_pagina.iter_rows(min_row=2, max_row=5):
    print(f'{linhas[0].value}\t{linhas[1].value}\t{linhas[2].value}')

#alterando o valor da celula
for linhas in frutas_pagina.iter_rows(min_row=2, max_row=5):
    for celula in linhas:
        if celula.value == 'Bananas':
            celula.value = 'Uva'

#Salvar as alterações
livro.save('Planilha de compras.xlsx')
